import pandas as pd
import os
import logging
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule

def write_report(df, output_path):
    """
    Writes the given DataFrame to an Excel file specified by output_path.
    Includes formatting and styling for better readability.
    """
    try:
        # Ensure the directory exists
        directory = os.path.dirname(output_path)
        if directory and not os.path.exists(directory):
            os.makedirs(directory)
        
        # Write to Excel with styling
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            # Write the main data
            df.to_excel(writer, index=False, sheet_name='Análisis de Retención')
            
            # Get the worksheet
            worksheet = writer.sheets['Análisis de Retención']
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
            
            # Format header row
            for cell in worksheet[1]:
                cell.style = 'Headline 1'
                
            # Add conditional formatting for continuity columns
            for col in df.columns:
                if col.startswith('Continuidad en'):
                    col_letter = worksheet[1][df.columns.get_loc(col)].column_letter
                    last_row = len(df) + 1
                    
                    # Add green fill for TRUE values
                    green_rule = CellIsRule(
                        operator='equal',
                        formula=['TRUE'],
                        stopIfTrue=True,
                        fill=PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
                    )
                    
                    # Add red fill for FALSE values
                    red_rule = CellIsRule(
                        operator='equal',
                        formula=['FALSE'],
                        stopIfTrue=True,
                        fill=PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
                    )
                    
                    # Apply rules to the column
                    worksheet.conditional_formatting.add(
                        f'{col_letter}2:{col_letter}{last_row}',
                        green_rule
                    )
                    worksheet.conditional_formatting.add(
                        f'{col_letter}2:{col_letter}{last_row}',
                        red_rule
                    )
            
            # Format percentage columns
            if 'Porcentaje de Continuidad' in df.columns:
                col_letter = worksheet[1][df.columns.get_loc('Porcentaje de Continuidad')].column_letter
                for row in range(2, len(df) + 2):
                    cell = worksheet[f'{col_letter}{row}']
                    cell.number_format = '0.00%'
        
        logging.info(f"Reporte generado exitosamente en: {output_path}")
        
    except Exception as e:
        logging.error(f"Error al escribir el reporte: {e}")
        raise
